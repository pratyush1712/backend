import openpyxl
import pandas as pd
import xlrd
from bs4 import BeautifulSoup
from selenium import webdriver
import time
from openpyxl import load_workbook
import constants


class Scraper:
    """ This class contains the backend code of the method."""

    def __init__(self, file, sheet, address, column_to_extract, column_to_insert):
        """ This method initializes the instance variables.
        It is used to link the frontend inputs to the backend."""
        self.filePath = file
        self.sheetName = sheet
        self.column_to_append = column_to_insert
        self.column_to_extract = column_to_extract
        # Dictionary with key as the name of item and value as the price.
        self.names = {}
        self.items = []
        self.prices = []
        self.all_website_items = []

    def extract_from_excel(self):
        """ This method extracts data from the given excel file."""
        ps = openpyxl.load_workbook(self.filePath)
        sheet = ps[self.sheetName]
        for row in range(1, sheet.max_row + 1):
            # each row in the spreadsheet represents information for a particular purchase.
            item = sheet[self.column_to_extract + str(row)].value
            self.names[item] = ""

    def parse_website(self, website):
        # IMPORTANT!!: Need to find a way to find location of chromedriver
        self.driver = webdriver.Chrome(constants.chromedriver_location)
        self.driver.get(website)
        time.sleep(5)
        self.html = self.driver.page_source
        self.soup = BeautifulSoup(self.html, "html.parser")

    def put_in_excel(self, data):
        b = "Col_" + self.column_to_append
        df_new = pd.DataFrame({b: data})
        wb = load_workbook(self.filePath)
        ws = wb[self.sheetName]
        a = self.column_to_append + "%d"
        for index, row in df_new.iterrows():
            cell = a % (index + 2)
            ws[cell] = row[0]

        wb.save(self.filePath)
